import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

def create_report(discrepancies, merged_df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Отчет"

    # Настройка стиля заголовков
    header_font = Font(bold=True, color="000000")
    header_alignment = Alignment(horizontal="center", vertical="center")

    # Устанавливаем заголовки и объединяем ячейки
    ws.merge_cells('B2:G2')
    ws['B2'].value = "Данные ЕКП"
    ws['B2'].font = header_font
    ws['B2'].alignment = header_alignment

    ws.merge_cells('B3:C3')
    ws['B3'].value = "Входящий остаток"
    ws['B3'].font = header_font
    ws['B3'].alignment = header_alignment

    ws.merge_cells('D3:E3')
    ws['D3'].value = "Обороты"
    ws['D3'].font = header_font
    ws['D3'].alignment = header_alignment

    ws.merge_cells('F3:G3')
    ws['F3'].value = "Исходящий остаток"
    ws['F3'].font = header_font
    ws['F3'].alignment = header_alignment

    ws['B4'].value = "Дт"
    ws['B4'].font = header_font
    ws['B4'].alignment = header_alignment

    ws['C4'].value = "Кт"
    ws['C4'].font = header_font
    ws['C4'].alignment = header_alignment

    ws['D4'].value = "Дт"
    ws['D4'].font = header_font
    ws['D4'].alignment = header_alignment

    ws['E4'].value = "Кт"
    ws['E4'].font = header_font
    ws['E4'].alignment = header_alignment

    ws['F4'].value = "Дт"
    ws['F4'].font = header_font
    ws['F4'].alignment = header_alignment

    ws['G4'].value = "Кт"
    ws['G4'].font = header_font
    ws['G4'].alignment = header_alignment

    ws.merge_cells('H2:M2')
    ws['H2'].value = "Данные ГБК"
    ws['H2'].font = header_font
    ws['H2'].alignment = header_alignment

    ws.merge_cells('H3:I3')
    ws['H3'].value = "Входящий остаток"
    ws['H3'].font = header_font
    ws['H3'].alignment = header_alignment

    ws.merge_cells('J3:K3')
    ws['J3'].value = "Обороты"
    ws['J3'].font = header_font
    ws['J3'].alignment = header_alignment

    ws.merge_cells('L3:M3')
    ws['L3'].value = "Исходящий остаток"
    ws['L3'].font = header_font
    ws['L3'].alignment = header_alignment

    ws['H4'].value = "Дт"
    ws['H4'].font = header_font
    ws['H4'].alignment = header_alignment

    ws['I4'].value = "Кт"
    ws['I4'].font = header_font
    ws['I4'].alignment = header_alignment

    ws['J4'].value = "Дт"
    ws['J4'].font = header_font
    ws['J4'].alignment = header_alignment

    ws['K4'].value = "Кт"
    ws['K4'].font = header_font
    ws['K4'].alignment = header_alignment

    ws['L4'].value = "Дт"
    ws['L4'].font = header_font
    ws['L4'].alignment = header_alignment

    ws['M4'].value = "Кт"
    ws['M4'].font = header_font
    ws['M4'].alignment = header_alignment

    ws.merge_cells('N2:S2')
    ws['N2'].value = "Расхождения"
    ws['N2'].font = header_font
    ws['N2'].alignment = header_alignment

    ws.merge_cells('N3:O3')
    ws['N3'].value = "Входящий остаток"
    ws['N3'].font = header_font
    ws['N3'].alignment = header_alignment

    ws.merge_cells('P3:Q3')
    ws['P3'].value = "Обороты"
    ws['P3'].font = header_font
    ws['P3'].alignment = header_alignment

    ws.merge_cells('R3:S3')
    ws['R3'].value = "Исходящий остаток"
    ws['R3'].font = header_font
    ws['R3'].alignment = header_alignment

    ws['N4'].value = "Дт"
    ws['N4'].font = header_font
    ws['N4'].alignment = header_alignment

    ws['O4'].value = "Кт"
    ws['O4'].font = header_font
    ws['O4'].alignment = header_alignment

    ws['P4'].value = "Дт"
    ws['P4'].font = header_font
    ws['P4'].alignment = header_alignment

    ws['Q4'].value = "Кт"
    ws['Q4'].font = header_font
    ws['Q4'].alignment = header_alignment

    ws['R4'].value = "Дт"
    ws['R4'].font = header_font
    ws['R4'].alignment = header_alignment

    ws['S4'].value = "Кт"
    ws['S4'].font = header_font
    ws['S4'].alignment = header_alignment

    ws.merge_cells('A2:A4')
    ws['A2'].value = 'Счет'
    ws['A2'].font = header_font
    ws['A2'].alignment = header_alignment  

    row_num = 5
    for index, row in merged_df.iterrows():
        ws.cell(row=row_num, column=1).value = row['№ Счета']
        ws.cell(row=row_num, column=2).value = row['Входящий отрицательный']
        ws.cell(row=row_num, column=3).value = row['Входящий положительный']
        ws.cell(row=row_num, column=4).value = row['Дебетовый оборот']
        ws.cell(row=row_num, column=5).value = row['Кредитовый оборот']
        ws.cell(row=row_num, column=6).value = row['Исходящий отрицательный']
        ws.cell(row=row_num, column=7).value = row['Исходящий положительный']
        ws.cell(row=row_num, column=8).value = row['в функциональной валюте']
        ws.cell(row=row_num, column=9).value = row['в функциональной валюте_1']
        ws.cell(row=row_num, column=10).value = row['в функциональной валюте_2']
        ws.cell(row=row_num, column=11).value = row['в функциональной валюте_3']
        ws.cell(row=row_num, column=12).value = row['в функциональной валюте_4']
        ws.cell(row=row_num, column=13).value = row['в функциональной валюте_5']
        ws.cell(row=row_num, column=14).value = row['difference_2_8']
        ws.cell(row=row_num, column=15).value = row['difference_3_9']
        ws.cell(row=row_num, column=16).value = row['difference_4_10']
        ws.cell(row=row_num, column=17).value = row['difference_5_11']
        ws.cell(row=row_num, column=18).value = row['difference_6_12']
        ws.cell(row=row_num, column=19).value = row['difference_7_13']  
        
     # Применение стиля к числовым ячейкам
        for col in range(1, 14):  # Столбцы с 1 по 13
            cell = ws.cell(row=row_num, column=col)
            if col == 1:  # Первая колонка (счет) без знаков после запятой
                cell.number_format = '0'
            else:  # Остальные числовые колонки с двумя знаками после запятой
                cell.number_format = '0.00'
        row_num += 1

     # Добавление строки "Итого:" и сумм по модулю для колонок 14-19
    total_row = row_num + 2
    ws.cell(row=total_row, column=1).value = "Итого:"
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    for col in range(14, 20):  # Колонки 14-19
        col_letter = chr(64 + col)  # Преобразование номера колонки в букву
        sum_formula = f"=SUMPRODUCT(ABS({col_letter}5:{col_letter}{row_num-1}))"
        ws.cell(row=total_row, column=col).value = sum_formula
        ws.cell(row=total_row, column=col).number_format = '0.00'
        ws.cell(row=total_row, column=col).font = Font(bold=True)

    # Добавление границ для всех ячеек с данными
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
    for row in ws.iter_rows(min_row=2, max_row=row_num, min_col=1, max_col=19):
        for cell in row:
            cell.border = thin_border

    return wb

def save_report(wb, save_path):
    # Сохранение файла
    wb.save(save_path)
